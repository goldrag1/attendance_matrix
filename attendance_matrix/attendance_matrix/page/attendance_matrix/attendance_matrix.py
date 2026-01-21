import frappe
from frappe import _
import json
from frappe.utils import getdate, nowdate, add_days, get_first_day, get_last_day, date_diff
# Import License Validator
from attendance_matrix.license import validate_license_hook

def get_permitted_departments():
    """
    Returns list of department names the current user has access to.
    Returns None if user has full access (System Manager, HR Manager only).
    HR User and other roles must use Department User Permission.
    """
    # Check if user has full access via Role (Only System Manager and HR Manager)
    user_roles = frappe.get_roles(frappe.session.user)
    if any(role in user_roles for role in ["System Manager", "HR Manager"]):
        return None  # Full access
    
    # Get departments from User Permission
    permitted = frappe.db.get_all(
        "User Permission",
        filters={
            "user": frappe.session.user,
            "allow": "Department"
        },
        pluck="for_value"
    )
    
    return permitted if permitted else None  # Return None if no restrictions set

@frappe.whitelist()
def get_matrix_data(month=None, year=None, department=None, company=None, employee=None, shift=None, overtime_type=None, active_only=True):
    # SECURITY: Enforce License Check (Cannot be bypassed by disabling hooks)
    validate_license_hook()

    # Ensure Custom Field for Dual Status exists
    make_custom_fields()

    # 0. Fetch Settings
    settings = frappe.get_single("Attendance Matrix Settings")
    status_map = settings.status_map
    
    # Manually serialize shift_map to ensure time is string (HH:mm:ss) not timedelta
    shift_map = []
    for s in settings.shift_map:
        shift_map.append({
            "shift_name": s.shift_name,
            "start_time": str(s.start_time) if s.start_time else "",
            "end_time": str(s.end_time) if s.end_time else ""
        })
    if not month or not year:
        today = getdate()
        month = today.month
        year = today.year
    
    first_day = get_first_day(f"{year}-{month}-01")
    last_day = get_last_day(first_day)
    
    # 1. Fetch Employees with Permission Filter
    filters = {}
    
    # Handle active_only filter
    # Check if active_only is "true" string (from JS) or True boolean
    is_active_only = str(active_only).lower() == "true"
    
    if is_active_only:
        filters['status'] = 'Active'

    # PERMISSION: Filter by User Permission
    permitted_depts = get_permitted_departments()
    if permitted_depts is not None:
        # User has department restrictions
        if department:
            if department in permitted_depts:
                filters['department'] = department
            else:
                frappe.throw(_("Bạn không có quyền truy cập phòng ban này / You don't have permission for this department"))
        else:
            filters['department'] = ['in', permitted_depts]
    elif department:
        # No restrictions, apply user's filter
        filters['department'] = department
    
    if company:
        filters['company'] = company
    if employee:
        # SEARCH LOGIC: Filter by ID OR Name
        # We need to drop 'name' from filters and use or_filters kwarg in get_all
        # But wait, get_all 'filters' param is AND. 
        # Frappe's get_all supports 'or_filters' as a separate argument.
        pass # Handle below in get_all call
    if shift:
        filters['default_shift'] = shift

    or_filters = None
    if employee:
        or_filters = {
            "name": ["like", f"%{employee}%"],
            "employee_name": ["like", f"%{employee}%"]
        }

    employees = frappe.get_all("Employee", 
        fields=["name", "employee_name", "department", "designation", "default_shift", "status"],
        filters=filters,
        or_filters=or_filters,
        order_by="employee_name asc"
    )

    # frappe.msgprint(f"DEBUG: Trả về {len(employees)} nhân viên. Meta: {first_day} - {last_day}") 
    
    # 2. Fetch Holidays
    holidays = []
    # Try to find default holiday list from Company or first employee
    holiday_list_name = None
    if employees:
        holiday_list_name = frappe.db.get_value("Employee", employees[0].name, "holiday_list")
    
    if not holiday_list_name and company:
        holiday_list_name = frappe.db.get_value("Company", company, "default_holiday_list")
        
    if holiday_list_name:
        holidays = frappe.get_all("Holiday",
            fields=["holiday_date", "description"],
            filters={"parent": holiday_list_name, "holiday_date": ["between", [first_day, last_day]]}

        )

    # 2.5 Filter Employees by Overtime Type (if selected)
    # This must be done BEFORE fetching full attendance to save performance, or AFTER?
    # Logic: If overtime_type is set, find all employees who have 'Attendance Overtime Log' with this type in the date range.
    if overtime_type:
        # 1. Find Attendance records in range
        sub_dates = frappe.get_all("Attendance", filters={
            "attendance_date": ["between", [first_day, last_day]],
            "docstatus": ["!=", 2],
            "company": company
        }, fields=["name", "employee"])
        
        att_ids = [d.name for d in sub_dates]
        
        if att_ids:
            # 2. Find Logs with specific type
            matching_logs = frappe.get_all("Attendance Overtime Log", filters={
                "parent": ["in", att_ids],
                "overtime_type": overtime_type
            }, fields=["parent"])
            
            # 3. Get Parent Attendance Names
            matching_att_names = [l.parent for l in matching_logs]
            
            # 4. Get Employee IDs from Attendance
            valid_employees = set([d.employee for d in sub_dates if d.name in matching_att_names])
            
            # 5. Filter the main employees list
            employees = [e for e in employees if e.name in valid_employees]
        else:
            employees = []

    # 3. Fetch Attendance
    # Fetch custom_matrix_status alongside standard status
    att_filters = {
        "attendance_date": ["between", [first_day, last_day]],
        "docstatus": ["!=", 2]
    }
    if department:
        att_filters["employee"] = ["in", [e.name for e in employees]]
        
    attendance_list = frappe.get_all("Attendance", 
        filters=att_filters,
        fields=["name", "employee", "attendance_date", "status", "working_hours", "custom_matrix_status"]
    )
    
    attendance = {}
    attendance_ids = []
    for a in attendance_list:
        key = f"{a.employee}_{a.attendance_date}"
        # PREFER Custom Status (User's Term) over ERPNext Status (Payroll Term) for Display
        display_status = a.custom_matrix_status if a.custom_matrix_status else a.status
        attendance[key] = {
            "status": display_status,
            "hours": a.working_hours,
            # detailed info for debugging if needed
            "real_status": a.status,
            "overtime": [] # Init
        }
        attendance_ids.append(a.name)

    # 3.1 Fetch Overtime Logs
    if attendance_ids:
        ot_logs = frappe.get_all("Attendance Overtime Log",
            fields=["parent", "overtime_type", "hours"],
            filters={"parent": ["in", attendance_ids]}
        )
        
        # Attach to attendance dict
        # We need to map parent (Attendance Name) to the Key (Employee_Date)
        # So we need a map: Attendance Name -> Key
        att_name_to_key = {a.name: f"{a.employee}_{a.attendance_date}" for a in attendance_list}
        
        for log in ot_logs:
            key = att_name_to_key.get(log.parent)
            if key and key in attendance:
                # Add to overtime list
                attendance[key]["overtime"].append({
                    "type": log.overtime_type,
                    "hours": log.hours
                })

    # 4. Fetch Meta for Filters
    all_companies = frappe.get_all("Company", fields=["name"], order_by="name asc")
    all_departments = frappe.get_all("Department", fields=["name", "company"], order_by="name asc")
    fiscal_years = frappe.get_all("Fiscal Year", fields=["name"], order_by="year_start_date desc")

    # Filter departments and companies based on user permission
    if permitted_depts is not None:
        # User has restrictions - only show permitted departments
        filtered_departments = [d.name for d in all_departments if d.name in permitted_depts]
        # Get companies that contain permitted departments
        permitted_companies = set(d.company for d in all_departments if d.name in permitted_depts and d.company)
        filtered_companies = [c.name for c in all_companies if c.name in permitted_companies]
        # Get employee names for autocomplete (already filtered in employees list)
        employee_options = [{"value": e["name"], "label": f"{e['employee_name']} ({e['name']})"} for e in employees]
    else:
        # User has full access - show all
        filtered_departments = [d.name for d in all_departments]
        filtered_companies = [c.name for c in all_companies]
        employee_options = []  # Empty = use search mode instead of dropdown

    # Get user permission info for display (use same logic as get_permitted_departments)
    user_roles = frappe.get_roles(frappe.session.user)
    has_full_access = any(role in user_roles for role in ["System Manager", "HR Manager"])
    
    permission_info = {
        "has_full_access": has_full_access,
        "permitted_departments": permitted_depts if permitted_depts else None,
        "user_display": frappe.session.user
    }

    return {
        "employees": employees,
        "attendance": attendance,
        "holidays": holidays,
        "meta": {
            "first_day": str(first_day),
            "last_day": str(last_day),
            "days_in_month": date_diff(last_day, first_day) + 1,
            "filter_options": {
                "companies": filtered_companies,
                "departments": filtered_departments,
                "fiscal_years": [f.name for f in fiscal_years],
                "employees": employee_options
            }
        },
        "settings": {
            "status_map": status_map,
            "shift_map": shift_map,
            "overtime_types": settings.overtime_types
        },
        "permission_info": permission_info
    }

@frappe.whitelist()
def save_matrix_bulk(data, mode="attendance"):
    # SECURITY: Enforce License Check
    validate_license_hook()

    """
    data = [
        { "employee": "EMP01", "date": "2026-01-01", "status": "..." }, # status here is the custom matrix status
        ...
    ]
    """
    changes = json.loads(data)
    
    settings = frappe.get_single("Attendance Matrix Settings")
    status_map_config = {s.status: s.payroll_status for s in settings.status_map} # Map: "Full ca" -> "Present"

    results = {"success": [], "errors": []}
    
    # DEBUG LOGGING (Temporary for silent save investigation)
    # frappe.log_error(f"Matrix Save Data: {data}, Mode: {mode}", "Attendance Matrix Save Debug")
    
    for item in changes:
        try:
            employee = item.get("employee")
            date = item.get("date")
            raw_status = item.get("status") # This is "Full ca" or "1"
            
            if not employee or not date: 
                continue
            
            # PERMISSION: Check if user can edit this employee
            permitted_depts = get_permitted_departments()
            if permitted_depts is not None:
                emp_dept = frappe.db.get_value("Employee", employee, "department")
                if emp_dept not in permitted_depts:
                    results["errors"].append({
                        "employee": employee,
                        "date": date,
                        "error": _("Bạn không có quyền chỉnh sửa nhân viên này / You don't have permission to edit this employee")
                    })
                    continue

            # Check if exists
            existing = frappe.db.exists("Attendance", {
                "employee": employee,
                "attendance_date": date,
                "docstatus": ["<", 2]
            })

            if mode == "attendance":
                # CASE 1: Deletion (Empty Status)
                if not raw_status:
                    if existing:
                        doc = frappe.get_doc("Attendance", existing)
                        if doc.docstatus == 1:
                            doc.cancel()
                        doc.delete()
                        results["success"].append(f"Deleted {employee} on {date}")
                    continue # Skip creation
                
                # CASE 2: Creation / Update
                # LOGIC: Dual Status Mapping
                final_custom_status = raw_status 
                final_erp_status = raw_status # Default to same if no mapping found

                # Build list of valid abbreviations
                valid_abbreviations = [s.abbreviation for s in settings.status_map if s.abbreviation]
                valid_statuses = [s.status for s in settings.status_map]
                all_valid = valid_abbreviations + valid_statuses

                # VALIDATION: Check if status is valid
                if raw_status not in all_valid:
                    valid_list = ", ".join(valid_abbreviations) if valid_abbreviations else ", ".join(valid_statuses[:5])
                    results["errors"].append({
                        "employee": employee,
                        "date": date,
                        "error": _(f"Trạng thái '{raw_status}' không hợp lệ. Các giá trị hợp lệ: {valid_list}")
                    })
                    continue

                if raw_status in status_map_config and status_map_config[raw_status]:
                    final_erp_status = status_map_config[raw_status]
                
                # SMART OPTION ADD (Fallback for unmapped custom statuses)
                if final_erp_status:
                     ensure_attendance_option(final_erp_status)
                
                # Handle Update vs Re-create
                should_create_new = True 
                
                # Check for existing data to preserve
                preserved_ot_logs = []
                preserved_ot_hours = 0.0

                if existing:
                    doc = frappe.get_doc("Attendance", existing)
                    
                    # Capture existing OT data
                    if doc.get("matrix_overtime_logs"):
                        for ot in doc.matrix_overtime_logs:
                            preserved_ot_logs.append({
                                "overtime_type": ot.overtime_type,
                                "hours": ot.hours
                            })
                    if doc.get("custom_overtime_hours"):
                        preserved_ot_hours = doc.custom_overtime_hours

                    if doc.docstatus == 1:
                        # If Submitted: Cancel & Delete -> Re-create
                        doc.cancel()
                        doc.delete()
                        should_create_new = True
                    else:
                        # If Draft: Update in place
                        doc.status = final_erp_status
                        doc.custom_matrix_status = final_custom_status
                        doc.flags.ignore_validate = True 
                        doc.save(ignore_permissions=True)
                        results["success"].append(f"Updated {employee} on {date}")
                        should_create_new = False

                if should_create_new:
                    doc = frappe.get_doc({
                        "doctype": "Attendance",
                        "employee": employee,
                        "attendance_date": date,
                        "status": final_erp_status,
                        "custom_matrix_status": final_custom_status,
                        "matrix_overtime_logs": preserved_ot_logs,
                        "custom_overtime_hours": preserved_ot_hours,
                        "docstatus": 1
                    })
                    # CRITICAL: Bypass hardcoded validation in Attendance.py (validate_status)
                    doc.flags.ignore_validate = True
                    doc.insert(ignore_permissions=True)
                    doc.submit() # Auto-Submit
                    results["success"].append(f"Created {employee} on {date}")
                    
            elif mode == "overtime":
                # OVERTIME LOGIC
                # 1. Parse raw_status (which is overtime string e.g. "OT1: 2; OT2: 1")
                # We need Overtime Types config
                ot_types_map = {ot.abbreviation: ot.overtime_name for ot in settings.overtime_types}
                
                parsed_entries = []
                total_hours = 0.0

                if raw_status:
                    # simplistic parser: "OT1: 2; OT2: 3"
                    parts = raw_status.split(';')
                    for part in parts:
                        if ':' in part:
                            abbr, qty = part.split(':', 1)
                            abbr = abbr.strip()
                            try:
                                qty = float(qty.strip())
                            except:
                                qty = 0
                            
                            type_name = ot_types_map.get(abbr)
                            if type_name:
                                parsed_entries.append({"overtime_type": type_name, "hours": qty})
                                total_hours += qty
                        else:
                            # Handle single value case if only 1 type exists? 
                            # Or if user just typed "2" and we assume default?
                            # For now, strict syntax or simplistic fallback?
                            # Design says: "Viet tat kieu lam them 1: so luong 1"
                            # Let's support just "number" if only 1 OT type is configured
                            if len(settings.overtime_types) == 1:
                                try:
                                    qty = float(part.strip())
                                    parsed_entries.append({"overtime_type": settings.overtime_types[0].overtime_name, "hours": qty})
                                    total_hours += qty
                                except:
                                    pass
                
                if existing:
                    doc = frappe.get_doc("Attendance", existing)
                    
                    # Handle Submitted Document: Cancel first
                    if doc.docstatus == 1:
                        doc.cancel()
                        # After cancel, we can either delete and recreate, or amend. 
                        # Simplest here for matrix mode is delete and recreate to ensure clean state
                        doc.delete()
                        # Re-create below logic
                        
                        doc = frappe.get_doc({
                            "doctype": "Attendance",
                            "employee": employee,
                            "attendance_date": date,
                            "status": doc.status, # Keep original status
                            "custom_matrix_status": doc.custom_matrix_status,
                            "matrix_overtime_logs": parsed_entries,
                            "custom_overtime_hours": total_hours, # New Total Field
                            "docstatus": 1
                        })
                        doc.flags.ignore_validate = True
                        doc.insert(ignore_permissions=True)
                        doc.submit()
                        results["success"].append(f"Updated (Re-created) Overtime for {employee} on {date}")
                    else:
                        # Draft: Update in place
                        doc.matrix_overtime_logs = []
                        for entry in parsed_entries:
                            doc.append("matrix_overtime_logs", entry)
                        
                        doc.custom_overtime_hours = total_hours # New Total Field

                        doc.flags.ignore_validate = True
                        doc.save(ignore_permissions=True)
                        # Auto-submit draft? User might want that.
                        doc.submit()
                        results["success"].append(f"Updated Overtime for {employee} on {date}")
                else:
                    # Auto-create Attendance if missing?
                    # Create as Present if OT exists
                    if parsed_entries:
                        doc = frappe.get_doc({
                            "doctype": "Attendance",
                            "employee": employee,
                            "attendance_date": date,
                            "status": "Present", # Default to Present
                            "matrix_overtime_logs": parsed_entries,
                            "custom_overtime_hours": total_hours, # New Total Field
                            "docstatus": 1
                        })
                        doc.flags.ignore_validate = True
                        doc.insert(ignore_permissions=True)
                        doc.submit() # Auto-submit
                        results["success"].append(f"Created Attendance (Present) with OT for {employee} on {date}")
                
        except Exception as e:
            frappe.log_error(f"Error saving matrix for {item}: {str(e)}")
            results["errors"].append(f"Error for {item.get('employee')} on {item.get('date')}: {str(e)}")
            
    return results

@frappe.whitelist()
def export_attendance_excel(month=None, year=None, department=None, company=None, employee=None, shift=None, abbreviation_mode=0, mode="attendance"):
    # SECURITY: Enforce License Check
    validate_license_hook()

    abbreviation_mode = int(abbreviation_mode)

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from frappe.utils import getdate, cstr

    # 1. Fetch Data
    data = get_matrix_data(month, year, department, company, employee, shift)
    meta = data['meta']
    data_map = data['attendance'] # { "Emp_Date": {status: "...", hours: ...} }
    
    # Settings for colors & calculations
    settings = frappe.get_single("Attendance Matrix Settings")
    
    # Map for Colors: Status -> Color
    # Also Map for Abbreviation if needed
    status_config_map = {} # Status -> {abbr, color}
    for s in settings.status_map:
        status_config_map[s.status] = {"abbr": s.abbreviation, "color": s.color}

    # Map for Summaries: Status -> Index in summary array
    # We want columns for each configured status
    if mode == "overtime":
        settings.overtime_types = settings.get("overtime_types", [])
        configured_statuses = [ot.overtime_name for ot in settings.overtime_types]
        summary_headers = [ot.overtime_name for ot in settings.overtime_types]
        ot_abbr_map = {ot.overtime_name: ot.abbreviation for ot in settings.overtime_types}

        # New Logic: Discover Legacy Types from Data
        legacy_types = set()
        for key, record in data_map.items():
             if record.get("overtime"):
                 for ot in record.get("overtime"):
                     t = ot.get("type")
                     if t and t not in configured_statuses:
                         legacy_types.add(t)
        
        # Append Legacy Types
        sorted_legacy = sorted(list(legacy_types))
        summary_headers.extend(sorted_legacy)
        for lt in sorted_legacy:
            ot_abbr_map[lt] = lt # Default abbr is Name
    else:
        configured_statuses = [s.status for s in settings.status_map]
        summary_headers = [s.status for s in settings.status_map]
        
        # New Logic: Discover Legacy Statuses in Data
        legacy_statuses = set()
        for key, record in data_map.items():
             if record.get("status"):
                 s = record.get("status")
                 if s and s not in configured_statuses:
                     legacy_statuses.add(s)
        
        # Append Legacy Statuses
        sorted_legacy = sorted(list(legacy_statuses))
        summary_headers.extend(sorted_legacy)
        # No map update needed as default is Name
    
    # 2. Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bang Cham Cong" if mode == "attendance" else "Bang Lam Them"

    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    sunday_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    saturday_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    today_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
    holiday_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") 

    # 3. Headers Construction
    # Fixed Headers
    # 3. Headers Construction
    # Fixed Headers
    headers = ["Mã NV", "Tên NV"]
    if mode == "overtime":
        headers.append("Kiểu tăng ca")
    else:
        headers.append("Ca")
    headers.append("Phòng ban")
    
    # Dynamic Date Headers
    first_day = getdate(meta['first_day'])
    days_in_month = meta['days_in_month']
    date_cols = []
    
    import datetime
    
    for i in range(days_in_month):
        d = first_day + datetime.timedelta(days=i)
        date_str = d.strftime("%Y-%m-%d")
        
        # We append the ACTUAL Date Object, not string, so Excel treats it as Date
        headers.append(d)
        
        date_cols.append({
            "date": date_str,
            "obj": d,
            "is_weekend": d.weekday() in [5, 6], 
            "is_sunday": d.weekday() == 6,
            "is_saturday": d.weekday() == 5
        })

    # Dynamic Summary Headers
    for status_label in summary_headers:
        headers.append(status_label)

    ws.append(headers)

    # 4. Filter & Sort Employees
    filters = {}
    if department: filters['department'] = department
    if company: filters['company'] = company
    if employee: filters['name'] = ["like", f"%{employee}%"]
    if shift: filters['default_shift'] = shift
    
    employees = frappe.get_all("Employee", 
        fields=["name", "employee_name", "department", "default_shift"], 
        filters=filters,
        order_by="employee_name asc"
    )

    holidays = data['holidays'] 
    holiday_dates = [h['holiday_date'] for h in holidays]

    current_date = getdate()
    is_current_month = current_date.month == first_day.month and current_date.year == first_day.year
    from frappe.utils import cstr

    # 5. Build Rows
    for emp in employees:
        # Determine Overtime Type value (like "Mixed" or specific type)
        ot_type_val = ""
        if mode == "overtime":
            # Scan all days for this employee to find dominant type
            found_types = set()
            for col_info in date_cols:
                date_str = col_info['date']
                key = f"{emp.name}_{date_str}"
                cell_data = data_map.get(key, {})
                if cell_data.get('overtime'):
                    for ot in cell_data['overtime']:
                         type_name = ot.get('type')
                         if type_name: found_types.add(type_name)
            
            if len(found_types) == 0:
                ot_type_val = ""
            elif len(found_types) == 1:
                ot_type_val = list(found_types)[0]
            else:
                ot_type_val = "Hỗn hợp"

        row = [emp.name, emp.employee_name]
        if mode == "overtime":
            row.append(ot_type_val)
        else:
            row.append(emp.default_shift)
        
        row.append(emp.department)
        
        # Init counters for this employee
        # Key: Status Name, Value: Count
        if mode == "overtime":
            summary_counts = {ot: 0.0 for ot in summary_headers}
        else:
            summary_counts = {s: 0 for s in summary_headers}

        # Date Values
        for col_info in date_cols:
            date_str = col_info['date']
            key = f"{emp.name}_{date_str}"
            
            cell_data = data_map.get(key, {})
            val = cell_data.get('status', '') # This is the Display Status (Full Name)
            
            # Count Logic (Before converting val to abbr)
            if mode == "attendance" and val:
                # Find matching config
                # Check exact match first
                if val in summary_counts:
                    summary_counts[val] += 1
                else:
                    # Check abbreviations (shouldn't happen if standardized but safe to keep)
                    for s in settings.status_map:
                        if s.abbreviation == val:
                            if s.status in summary_counts:
                                summary_counts[s.status] += 1
                            break
            
            elif mode == "overtime":
                # For summary we need specific overtime hours per type
                # cell_data['overtime'] is list of {type: ..., hours: ...}
                if cell_data.get("overtime"):
                     for ot_entry in cell_data["overtime"]:
                         ot_type = ot_entry.get("type")
                         ot_hours = ot_entry.get("hours") or 0.0
                         if ot_type in summary_counts:
                             summary_counts[ot_type] += ot_hours

            # ABBREVIATION MODE CHECK
            display_val = val
            if mode == "attendance" and abbreviation_mode == 1 and val:
                 if val in status_config_map and status_config_map[val]["abbr"]:
                     display_val = status_config_map[val]["abbr"]
            elif mode == "overtime":
                # Display overtime string like "OT1: 2" or just "2" if one type
                if cell_data.get("overtime"):
                     ot_parts = []
                     for ot_entry in cell_data["overtime"]:
                         ot_type = ot_entry.get("type")
                         ot_hours = ot_entry.get("hours")
                         
                         abbr = ot_abbr_map.get(ot_type, ot_type)
                         ot_parts.append(f"{abbr}: {ot_hours}")
                     display_val = "\n".join(ot_parts)
                else:
                    display_val = ""

            row.append(display_val)
        
        # Summary Values
        for s in summary_headers:
            row.append(summary_counts[s])

        ws.append(row)
        
        # --- STYLING THE ROW ---
        row_idx = ws.max_row
        max_col = len(headers)

        # Apply Border to ALL cells in row
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = thin_border
            
            # Center Align everything except maybe Name? Use Center for now for consistency
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Specific Styling for Date Columns
        # Date cols start at index 5 (1-based) -> 4 + 1
        # Length is len(date_cols)
        start_date_col_idx = 5
        
        for i, col_info in enumerate(date_cols):
            col_idx = start_date_col_idx + i
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value

            fill = None
            
            # P1: Status Color (Only for Attendance Mode?)
            # Or Overtime Mode should have indicators?
            # User wants standard look, maybe just keep holiday/weekend colors
            
            if mode == "attendance":
                date_str = col_info['date']
                key = f"{emp.name}_{date_str}"
                cell_data = data_map.get(key, {})
                original_status = cell_data.get('status') # This is the Full Name / Custom Status

                if original_status:
                    config = status_config_map.get(original_status)
                    if config and config.get("color"):
                        hex_color = config["color"]
                        hex_clean = hex_color.replace("#", "")
                        if len(hex_clean) == 6:
                            fill = PatternFill(start_color=hex_clean.upper(), end_color=hex_clean.upper(), fill_type="solid")
            
            # P2: Holiday
            if not fill and not val and col_info['date'] in holiday_dates:
                 fill = holiday_fill
            
            # P3: Today
            if not fill and not val and is_current_month and col_info['date'] == cstr(current_date):
                fill = today_fill

            # P4: Weekend
            if not fill and not val:
                if col_info['is_sunday']: fill = sunday_fill
                elif col_info['is_saturday']: fill = saturday_fill
            
            if fill:
                cell.fill = fill

        # Specific Styling for Summary Columns
        start_summary_col_idx = 5 + len(date_cols)
        for i, status_label in enumerate(configured_statuses):
            col_idx = start_summary_col_idx + i
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if mode == "attendance":
                config = status_config_map.get(status_label)
                if config and config.get("color"):
                    hex_color = config["color"]
                    hex_clean = hex_color.replace("#", "")
                    if len(hex_clean) == 6:
                        summary_fill = PatternFill(start_color=hex_clean.upper(), end_color=hex_clean.upper(), fill_type="solid")
                        cell.fill = summary_fill

    # --- HEADER STYLING & FORMATTING ---
    # We do this last or first, doesn't matter much, but let's do it cleanly
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format Date Columns (indices related to date_cols)
        # Date cols are from 5 to 5+len-1
        if 5 <= col_idx < 5 + len(date_cols):
            # Apply Excel Date Format: "d ddd" e.g. "1 Sun"
            # Or standard "dd/mm/yyyy" if user wants "chuẩn ngày tháng"
            cell.number_format = 'dd/mm/yyyy' 
            # Also apply header color for weekends/holidays in HEADER row too?
            # Let's match the date index
            date_idx = col_idx - 5
            d_info = date_cols[date_idx]
            
            if d_info['date'] in holiday_dates:
                cell.fill = holiday_fill
                cell.font = Font(bold=True, color="FF0000")
            elif d_info['is_sunday']:
                cell.fill = sunday_fill
            elif d_info['is_saturday']:
                cell.fill = saturday_fill
        
        # Summary Headers Styling (Bold, maybe different color?)
        if col_idx >= 5 + len(date_cols):
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Darker gray

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 15 # ID
    ws.column_dimensions['B'].width = 25 # Name
    ws.column_dimensions['C'].width = 15 # Shift
    ws.column_dimensions['D'].width = 20 # Dept
    
    # Date Cols Width
    for i in range(len(date_cols)):
        col_letter = openpyxl.utils.get_column_letter(5 + i)
        ws.column_dimensions[col_letter].width = 12 # Wider for dd/mm/yyyy

    # Summary Cols Width
    for i in range(len(configured_statuses)):
        col_letter = openpyxl.utils.get_column_letter(5 + len(date_cols) + i)
        ws.column_dimensions[col_letter].width = 10

    # Save
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response.filename = f"ChamCong_{month}_{year}{'_OT' if mode == 'overtime' else ''}.xlsx"
    frappe.response.filecontent = output.getvalue()
    frappe.response.type = "binary"

def make_custom_fields():
    custom_fields = {
        "Attendance": [
            # 1. Custom Status for Matrix
            {
                "fieldname": "custom_matrix_status",
                "label": "Matrix Display Status",
                "fieldtype": "Data",
                "insert_after": "status",
                "hidden": 1
            },
            # 2. Overtime Section & Table (If not already created by patch)
            {
                "fieldname": "matrix_overtime_section",
                "fieldtype": "Section Break",
                "label": "Matrix Overtime Logs",
                "insert_after": "early_exit" 
            },
            {
                "fieldname": "matrix_overtime_logs",
                "fieldtype": "Table",
                "label": "Overtime Logs (Matrix)",
                "options": "Attendance Overtime Log",
                "insert_after": "matrix_overtime_section"
            },
            # 3. TOTAL Overtime Hours (New)
            {
                "fieldname": "custom_overtime_hours",
                "fieldtype": "Float",
                "label": "Tổng số lượng / giờ",
                "insert_after": "matrix_overtime_logs",
                "read_only": 1 
            }
        ]
    }

    from frappe.custom.doctype.custom_field.custom_field import create_custom_fields
    create_custom_fields(custom_fields, update=False) # update=False to avoid re-creation cost, only missing ones

def ensure_attendance_option(status_val):
    meta = frappe.get_meta("Attendance")
    field = meta.get_field("status")
    current_options = (field.options or "").split("\n")
    
    if status_val not in current_options:
        # Auto-add
        current_options.append(status_val)
        updated_options_str = "\n".join(current_options)
        frappe.make_property_setter({
            "doctype": "Attendance",
            "fieldname": "status",
            "property": "options",
            "value": updated_options_str,
            "doctype_or_field": "DocField"
        })
        frappe.db.commit() 
        frappe.clear_cache(doctype="Attendance")
        frappe.get_meta("Attendance", cached=False)


@frappe.whitelist()
def save_attendance_settings(status_map, shift_map, overtime_types=None):
    doc = frappe.get_single("Attendance Matrix Settings")
    
    if isinstance(status_map, str):
        status_map = json.loads(status_map)
    if isinstance(shift_map, str):
        shift_map = json.loads(shift_map)
    
    overtime_types_list = []
    if overtime_types:
        if isinstance(overtime_types, str):
            overtime_types_list = json.loads(overtime_types)
        else:
            overtime_types_list = overtime_types

    # --- Force Clean Strategy (The only way to kill duplicates reliably) ---
    
    def sync_force_clean(doc_list_name, input_list, key_field):
        # 1. Deduplicate Input List
        unique_inputs = []
        seen_keys = set()
        
        for row in input_list:
            k = row.get(key_field)
            if k and k not in seen_keys:
                seen_keys.add(k)
                unique_inputs.append(row)
            elif not k:
                pass
                
        # 2. CLEAR ALL existing rows
        doc.set(doc_list_name, [])
        
        # 3. Insert Unique Rows with Explicit Index
        for i, row in enumerate(unique_inputs):
            # Explicitly set idx to ensure order is persisted
            row['idx'] = i + 1
            doc.append(doc_list_name, row)

    # Execute Sync
    sync_force_clean("status_map", status_map, "abbreviation")
    sync_force_clean("shift_map", shift_map, "shift_name")
    if overtime_types_list:
        sync_force_clean("overtime_types", overtime_types_list, "abbreviation")

    # --- SMART SCHEMA UPDATE: Auto-add Status options to Attendance DocType ---
    try:
        # 1. Get current options
        meta = frappe.get_meta("Attendance")
        field = meta.get_field("status")
        current_options = (field.options or "").split("\n")
        
        # 2. Collect unique statuses from new settings
        new_statuses = set()
        for s in status_map:
            if s.get("status"):
                new_statuses.add(s.get("status"))
        
        # 3. Check for missing options
        added_options = []
        for s in new_statuses:
            if s not in current_options:
                current_options.append(s)
                added_options.append(s)
        
        # 4. If changes needed, update via Property Setter
        if added_options:
            updated_options_str = "\n".join(current_options)
            
            # Use Property Setter to avoid editing source code
            frappe.make_property_setter({
                "doctype": "Attendance",
                "fieldname": "status",
                "property": "options",
                "value": updated_options_str,
                "doctype_or_field": "DocField"
            })
            # frappe.msgprint(f"Đã tự động thêm các trạng thái mới vào hệ thống: {', '.join(added_options)}")
            
            # Clear cache to apply immediately
            frappe.clear_cache(doctype="Attendance")
            
    except Exception as e:
        # Don't block saving if schema update fails (permission issues etc)
        frappe.log_error(f"Failed to update Attendance options: {str(e)}", "Attendance Matrix Error")

    doc.save()
    return "OK"

@frappe.whitelist()
def set_employee_shift(employee, shift_name):
    # SECURITY: Enforce License Check (Optional here but good hygiene)
    validate_license_hook()

    # 1. Check if Shift Type exists
    if not frappe.db.exists("Shift Type", shift_name):
        # 2. Try to find definition in Settings
        settings = frappe.get_single("Attendance Matrix Settings")
        found_config = None
        for s in settings.shift_map:
            if s.shift_name == shift_name:
                found_config = s
                break
        
        if found_config:
            # 3. Create Shift Type
            new_shift = frappe.get_doc({
                "doctype": "Shift Type",
                "name": shift_name,
                "start_time": str(found_config.start_time),
                "end_time": str(found_config.end_time)
            })
            new_shift.insert(ignore_permissions=True)
            frappe.msgprint(f"Đã tự động tạo loại ca mới: {shift_name}")
        else:
            frappe.throw(f"Ca '{shift_name}' chưa được định nghĩa trong Cấu hình Chấm công. Vui lòng thêm vào danh sách cấu hình trước.")

    # 4. Update Employee
    frappe.db.set_value("Employee", employee, "default_shift", shift_name)
    return "OK"
